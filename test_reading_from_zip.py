import zipfile
import pytest
from pypdf import PdfReader
import csv
from openpyxl import load_workbook


from conftest import TMP_DIR


def test_checking_archive_csv(create_archive):
    with zipfile.ZipFile(TMP_DIR+'/test.zip') as zip_file:
        with zip_file.open('TESTCSV.csv') as csv_file:
            content = csv_file.read().decode(
                'utf-8-sig')
            csvreader = list(csv.reader(content.splitlines(), delimiter=';'))
            second_row = csvreader[1]

            assert second_row[0] == 'OU002'
            assert second_row[1] == 'OU001'

def test_checking_archive_xlsx(create_archive):
    with zipfile.ZipFile(TMP_DIR+'/test.zip') as zip_file:
        with zip_file.open('TESTXLSX.xlsx') as xlsx_file:
            content = load_workbook(xlsx_file)
            sheet = content.active

            assert sheet.cell(row=3, column=2).value == "Сергеев"
            assert sum(1 for _ in sheet.iter_rows()) == 6
            assert sum(1 for _ in sheet.iter_cols()) == 13


def test_checking_archive_pdf(create_archive):
    with zipfile.ZipFile(TMP_DIR + '/test.zip') as zip_file:
        with zip_file.open('TESTPDF.pdf') as pdf_file:
            reader = PdfReader(pdf_file)

            assert len(reader.pages) == 1
            assert "Тестовый PDF-документ" in reader.pages[0].extract_text()